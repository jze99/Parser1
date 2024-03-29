from XMLparser import xml_parser
import openpyxl
import datetime

def uploadPath(file_path: str, directory_file:str):
    
    if file_path!="" and directory_file == "":
        workbook = openpyxl.load_workbook(file_path)
        if '1' not in workbook.sheetnames:
            workbook.create_sheet('1')
            sheet = workbook['1']
        else:
            sheet = workbook['1']
    
    if file_path == "" and directory_file != "":
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("1")
        workbook.create_sheet("2")
    
    newObjets = xml_parser.ObjectMass
    row_number = sheet.max_row
    
    sheet.cell(row=row_number,column=1).value = "Кадастровый номер"
    sheet.cell(row=row_number,column=2).value = "вид объекта (ОКС, ЗУ)"
    sheet.cell(row=row_number,column=3).value = "Номер точки"
    sheet.cell(row=row_number,column=4).value = "X"
    sheet.cell(row=row_number,column=5).value = "У"
    sheet.cell(row=row_number,column=6).value = "Погрешность"
    sheet.cell(row=row_number,column=7).value = "Метод определения точки"
    sheet.cell(row=row_number,column=8).value = "Источник"
    
    row_number+=1
    
    for iobj, number in enumerate(newObjets):
        if len(number.x) != 0:
            for xi, xobj in enumerate(number.x):
                sheet.cell(row=row_number, column=1).value = number.cadNumber
                sheet.cell(row=row_number, column=2).value = number.view
                sheet.cell(row=row_number, column=3).value = int(number.ord_nmb[xi])
                sheet.cell(row=row_number, column=4).value = float(xobj)
                sheet.cell(row=row_number, column=5).value = float(number.y[xi])
                if len(number.errorRate) != 0:
                    sheet.cell(row=row_number, column=6).value = number.errorRate[xi]
                sheet.cell(row=row_number, column=8).value = "КПТ"
                row_number+=1
                
    if '2' not in workbook.sheetnames:
        workbook.create_sheet('2')
        sheet = workbook['2']
    else:
        sheet = workbook['2']
    
    row_number = sheet.max_row
    
    sheet.cell(row=row_number,column=1).value = "Кадастровый номер"
    sheet.cell(row=row_number,column=2).value = "вид объекта (ОКС, ЗУ)"
    sheet.cell(row=row_number,column=3).value = "Площадь"
    sheet.cell(row=row_number,column=4).value = "Погрешность определения площади"
    sheet.cell(row=row_number,column=5).value = "Наличие координат"
    
    row_number+=1
    
    for iobj, number in enumerate(newObjets):
        sheet.cell(row=row_number, column=1).value = number.cadNumber
        sheet.cell(row=row_number, column=2).value = number.view
        sheet.cell(row=row_number, column=3).value = number.square
        sheet.cell(row=row_number, column=4).value = number.area_error
        if len(number.x) == 0:
            sheet.cell(row=row_number, column=5).value ="нет координат в ЕГРН"
        else:
            sheet.cell(row=row_number, column=5).value ="есть координаты в ЕГРН"
            
        row_number+=1
    try:
        
        if file_path!="" and directory_file == "":
            workbook.save(file_path)
        if file_path == "" and directory_file != "":
            current_datetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            file_name = f"/{current_datetime}.xlsx"
            workbook.save(directory_file+file_name)
        
    except PermissionError:
        return False
    
    finally:
        workbook.close()
        
    return True